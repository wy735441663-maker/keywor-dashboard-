import sys, json, os

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

OUTPUT_DIR = os.environ.get("EXCEL_OUTPUT_DIR", "C:/Users/Administrator/Desktop/文件夹/AI学习/try/项目名称")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Excel_progame.xlsx")

def parse_items(raw):
    if not raw or not str(raw).strip():
        return []
    text = str(raw)
    for ch in [",", chr(65292), chr(12290), ";", chr(65307)]:
        text = text.replace(ch, chr(10))
    items = [s.strip() for s in text.split(chr(10)) if s.strip()]
    return items

def main():
    input_file = sys.argv[1] if len(sys.argv) > 1 else None
    if input_file:
        with open(input_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = json.load(sys.stdin)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active

    from openpyxl.styles import Font, Alignment, PatternFill
    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    N = chr(10)

    headers = [chr(39033)+chr(30446)+chr(21517)+chr(31216), "ASIN", chr(20851)+chr(38190)+chr(35789), chr(36127)+chr(36131)+chr(20154)]
    ws.append(headers)
    for ci in range(1, 5):
        c = ws.cell(row=1, column=ci)
        c.font = hdr_font
        c.fill = hdr_fill

    total = 0
    for project in data:
        name = str(project.get("name", ""))
        owner = str(project.get("owner", ""))
        asins = project.get("asins") or [project.get("asin", "")]
        kw_raw = project.get("keywords", [])
        if isinstance(kw_raw, str):
            keywords = parse_items(kw_raw)
        else:
            keywords = kw_raw
        if not isinstance(asins, list):
            asins = parse_items(str(asins))
        asin_list = [str(a).strip() for a in asins if a and str(a).strip()]
        kw_list = [str(k).strip() for k in keywords if k and str(k).strip()]
        if not asin_list or not kw_list:
            continue
        ws.append([name, N.join(asin_list), N.join(kw_list), owner])
        total += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 12
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap
    for i in range(2, ws.max_row + 1):
        kw_cell = ws.cell(row=i, column=3)
        lines = kw_cell.value.count(N) + 1 if kw_cell.value else 1
        ws.row_dimensions[i].height = min(lines * 15, 300)

    final = OUTPUT_FILE
    try:
        wb.save(final)
    except PermissionError:
        import tempfile, shutil
        final = OUTPUT_FILE.replace(".xlsx", f"_{total}rows.xlsx")
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.close()
        wb.save(tmp.name)
        shutil.move(tmp.name, final)

    print(json.dumps({"success": True, "path": final, "rows": total, "projects": len(data)}))

if __name__ == "__main__":
    main()
